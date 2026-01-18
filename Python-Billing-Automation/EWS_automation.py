# EWS_automation.py
# FINAL CLIENT-WISE + FULL FORMATTING | 100% Perfect

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
from datetime import datetime
import tkinter as tk
from tkinter import messagebox

MASTER_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Master\Database FIleV1.xlsx"
EWS_OUTPUT_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\EWS_Report.xlsx"

SOURCE_SHEET = "Master_Database"
MONTHLY_SHEET = "Monthly_Master_Database"

# Fixed columns
FIXED_COLUMNS = {
    "Client_Code": 1, "Segment": 2, "Customer_AS_PER_SOW": 3, "Order_Title_AS_per_SOW": 4,
    "Sr_Manager": 5, "Lead": 6, "Type": 7, "Billing_Zone_Type": 8, "Billing_Type": 9,
    "Clarivate_POC": 10, "Client_POC": 11, "CSM_Contact": 12,
    "EWS_Indicator": 13, "RAG_Historic_Comments": 14, "RAG_Current_Brief_Summary": 15,
}

# Colors
BLUE_HEADER = "0070C0"
GREEN_HEADER = "92D050"
DARK_BLUE_HEADER = "002060"

RAG_COLORS = {"GREEN": "C6E0B4", "RED": "F4B084", "AMBER": "FFE699", "YELLOW": "FFE699"}

# Styles
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
white_bold = Font(color="FFFFFF", bold=True)
thin = Side(border_style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cal = None

def get_client_base(code):
    if not code or '-' not in code: return code
    parts = code.split('-')
    return f"{parts[0]}-{parts[1]}" if len(parts) >= 2 else code

def generate_ews_report():
    global cal
    try:
        if cal is None:
            messagebox.showerror("Error", "Select month first!")
            return

        selected_date = cal.get_date()
        current_year = selected_date.year
        current_month = selected_date.month

        # Last 12 months
        months_12 = []
        temp_y, temp_m = current_year, current_month
        for _ in range(12):
            months_12.append(datetime(temp_y, temp_m, 1).strftime("%b%y"))
            temp_m -= 1
            if temp_m == 0:
                temp_m = 12
                temp_y -= 1
        months_12 = months_12[::-1]

        wb = openpyxl.load_workbook(MASTER_PATH)
        ws_master = wb[SOURCE_SHEET]
        ws_monthly = wb[MONTHLY_SHEET]

        master_headers = {str(ws_master.cell(1,c).value or "").strip(): c for c in range(1, ws_master.max_column + 1)}

        # Monthly columns
        client_col = rag_col = month_col = None
        for c in range(1, ws_monthly.max_column + 1):
            h = str(ws_monthly.cell(1, c).value or "").strip()
            if "Client_Code" in h: client_col = c
            if "RAG_Status" in h: rag_col = c
            if "Billing_Month" in h: month_col = c

        if not all([client_col, rag_col, month_col]):
            messagebox.showerror("Error", "Monthly sheet missing columns!")
            return

        # RAG dictionary
        rag_dict = {}
        for r in range(2, ws_monthly.max_row + 1):
            client = str(ws_monthly.cell(r, client_col).value or "").strip()
            raw = str(ws_monthly.cell(r, month_col).value or "").strip()
            rag = str(ws_monthly.cell(r, rag_col).value or "").strip().upper()
            if not client or not raw: continue
            try:
                parts = raw.split("-")
                if len(parts) >= 2:
                    y, m = parts[0], parts[1]
                    mkey = m[:3].title() + y[-2:]
                    base = get_client_base(client)
                    if mkey in months_12:
                        rag_dict[(base, mkey)] = rag
            except: continue

        # Client-wise data
        client_data = {}
        for r in range(2, ws_master.max_row + 1):
            if str(ws_master.cell(r, master_headers.get("Billing_On_Off",1)).value or "").strip().lower() != "yes":
                continue
            full_code = str(ws_master.cell(r, master_headers["Client_Code"]).value or "").strip()
            base = get_client_base(full_code)
            if not base: continue

            if base not in client_data:
                client_data[base] = {"first_row": r, "rag": {m: "" for m in months_12}}

            # Update RAG
            for mkey in months_12:
                if (base, mkey) in rag_dict:
                    client_data[base]["rag"][mkey] = rag_dict[(base, mkey)]

        # Create Report
        wb_ews = openpyxl.Workbook()
        ws = wb_ews.active
        ws.title = "EWS_Report"

        # Headers
        for name, col in FIXED_COLUMNS.items():
            ws.cell(1, col).value = name

        rag_col_map = {}
        for i, mkey in enumerate(months_12):
            col = 16 + i
            ws.cell(1, col).value = f"{mkey}_RAG"
            rag_col_map[mkey] = col

        # Fill data
        row = 2
        for base, data in client_data.items():
            r = data["first_row"]
            for name, pos in FIXED_COLUMNS.items():
                src = master_headers.get(name)
                if src:
                    cell = ws.cell(row, pos)
                    cell.value = ws_master.cell(r, src).value
                    cell.alignment = center
                    cell.border = border

            for mkey in months_12:
                val = data["rag"][mkey]
                cell = ws.cell(row, rag_col_map[mkey])
                cell.value = val
                cell.alignment = center
                cell.border = border
                if val in RAG_COLORS:
                    cell.fill = PatternFill("solid", RAG_COLORS[val])

            row += 1

        # FULL FORMATTING
        blue_fill = PatternFill("solid", BLUE_HEADER)
        green_fill = PatternFill("solid", GREEN_HEADER)
        dark_fill = PatternFill("solid", DARK_BLUE_HEADER)

        for c in range(1, 28):
            cell = ws.cell(1, c)
            cell.font = white_bold
            cell.alignment = center
            cell.border = border
            if c <= 9: cell.fill = blue_fill
            elif c <= 15: cell.fill = green_fill
            else: cell.fill = dark_fill

        # Auto-fit
        for i in range(1, 28):
            col_letter = openpyxl.utils.get_column_letter(i)
            max_len = max((len(str(c.value or "")) for c in ws[col_letter]), default=10)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb_ews.save(EWS_OUTPUT_PATH)
        os.startfile(EWS_OUTPUT_PATH)
        messagebox.showinfo("Tool is Done and working file!",
                            f"Client-wise EWS Report Ready!\n"
                            f"Total Clients: {len(client_data)}\n"
                            f"Full Formatting Applied!")

    except Exception as e:
        messagebox.showerror("Error", str(e))