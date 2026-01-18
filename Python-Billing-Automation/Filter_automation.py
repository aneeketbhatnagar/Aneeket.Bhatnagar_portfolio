import openpyxl
import tkinter as tk
from tkinter import messagebox
from openpyxl.worksheet.filters import AutoFilter

MASTER_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Master\Database FIleV1.xlsx"

SOURCE_SHEET = "Master_Database"
TARGET_SHEET = "Filtered_Data"

# Column mapping:  { "ColumnName_in_Master": Position_in_New_Sheet }
COLUMN_ORDER = {
    "Client_Code": 1,
    "Manager": 2,
    "Sr_Manager": 33,
    "Billing_Entity": 3,
    "Lead": 4,
    "Customer_AS_PER_SOW": 5,
    "Order_Title_AS_per_SOW": 6,
    "Billing_Type": 7,
    "No._of_FTEs": 8,
    "Fee_per_FTE": 9,
    "Total_FTE_Fee": 10,
    "Hours_Per_FTE": 11,
    "Total_Committed_Hours": 12,
    "Actual_Hours": 13,
    "Overage_Hours": 14,
    "Overage_Rate": 15,
    "Total_Overages": 16,
    "Total_Amount": 17,
    "Ship_to_address": 18,
    "Ship_to_email": 19,
    "Ship_to_SAP": 20,
    "Bill_to_address": 21,
    "Bill_to_email": 22,
    "Bill_to_SAP": 23,
    "Narrative": 24,
    "Deal_Type": 25,
    "Docket_PO_No": 26,
    "Product_type": 27,
    "Campaign-Code": 28,
    "Comment_Or_Instructions": 29,
    "Currency": 30,
    "Billing_Status": 31,
    "Billing_Month":32,
    "BillingPocStatus":34
}

def create_filtered_sheet(mon_name):
    wb = openpyxl.load_workbook(MASTER_PATH)

    if SOURCE_SHEET not in wb.sheetnames:
        print(f"Sheet '{SOURCE_SHEET}' not found.")
        return

    ws = wb[SOURCE_SHEET]

    # Remove all sheet level filters
    #if wb.auto_filter:
    #    wb.auto_filter = AutoFilter()  # resets filters safely


    # Remove old output sheet if exists
    if TARGET_SHEET in wb.sheetnames:
        del wb[TARGET_SHEET]
    new_ws = wb.create_sheet(TARGET_SHEET)

    # Read header row
    header_map = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            header_map[str(name).strip()] = col

    # Create NEW header row
    for col_name, new_pos in COLUMN_ORDER.items():
        new_ws.cell(row=1, column=new_pos).value = col_name

    # Copy rows where Billing_On_Off = Yes
    target_row = 2

    if "Billing_On_Off" not in header_map:
        print("Column 'Billing_On_Off' not found.")
        return

    billing_col = header_map["Billing_On_Off"]

    for row in range(2, ws.max_row + 1):
        flag = ws.cell(row=row, column=billing_col).value
        if str(flag).strip().lower() != "yes":
            continue

        # Copy selected columns
        for col_name, new_pos in COLUMN_ORDER.items():
            if col_name in header_map:
                src_col = header_map[col_name]
                val = ws.cell(row=row, column=src_col).value
                new_ws.cell(row=target_row, column=new_pos).value = val

        # ---------- ADD FORMULAS ----------
        # Column 14 formula: =IF(M2 > L2, M2 - L2, "0")
        new_ws.cell(row=target_row, column=14).value = f'=IF(M{target_row}>L{target_row}, M{target_row}-L{target_row}, "0")'

        # Column 16 formula: =O2 * N2
        new_ws.cell(row=target_row, column=16).value = f'=O{target_row}*N{target_row}'

        # Column 31 Add Month
        new_ws.cell(row=target_row, column=32).value = f"{mon_name}"
        # -----------------------------------

        target_row += 1

    wb.save(MASTER_PATH)
    print("Filtered data created successfully with formulas.")
    messagebox.showinfo("Info", "Monthly File  created successfully")

