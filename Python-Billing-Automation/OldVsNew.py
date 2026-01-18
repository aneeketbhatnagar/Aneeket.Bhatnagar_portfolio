import os
import shutil
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.filters import AutoFilter

MASTER_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Master\Database FIleV1.xlsx"
RECEIVED_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Received\Processed"

SHEET_NAME = "Monthly"
LAST_COL_INDEX = 34  # compare columns 1 to 34


def ensure_backup(path: str) -> str:
    """Optional: create a backup copy of the master file."""
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(path)
    backup = f"{base}_backup_{ts}{ext}"
    shutil.copy2(path, backup)
    return backup


def get_row_values(ws, row: int):
    """Return list of values from col 1..LAST_COL_INDEX for a given row."""
    return [
        ws.cell(row=row, column=col_idx).value
        for col_idx in range(1, LAST_COL_INDEX + 1)
    ]


def run_old_vs_new():
    wb_master = load_workbook(MASTER_PATH, data_only=False)
    ws_master = wb_master[SHEET_NAME]

    # Remove all sheet level filters
    if ws_master.auto_filter:
        ws_master.auto_filter = AutoFilter()  # resets filters safely

    # ---------- Build master_index: Client_Code -> list of row numbers ----------
    master_index = {}
    for row_idx in range(2, ws_master.max_row + 1):
        val = ws_master.cell(row=row_idx, column=1).value  # Client_Code is column 1
        if val is None:
            continue
        key = str(val).strip()
        if key:
            master_index.setdefault(key, []).append(row_idx)

    # ---------- Create / reset Changed_Line_Items sheet ----------
    if "Changed_Line_Items" in wb_master.sheetnames:
        del wb_master["Changed_Line_Items"]

    ws_change = wb_master.create_sheet("Changed_Line_Items")

    # Set header row
    change_headers = [
        "Client_Code",
        "Column_No",
        "Column_Name",
        "Master_Value",
        "Received_Value",
    ]
    for col_idx, header in enumerate(change_headers, start=1):
        ws_change.cell(row=1, column=col_idx).value = header

    # ---------- Build header name map for master (row 1) ----------
    master_headers = {}
    for col_idx in range(1, LAST_COL_INDEX + 1):
        header_val = ws_master.cell(row=1, column=col_idx).value
        if header_val:
            master_headers[col_idx] = header_val

    # ---------- Loop over received files ----------
    for fname in os.listdir(RECEIVED_PATH):
        if not fname.lower().endswith((".xlsx", ".xlsm")):
            continue

        fpath = os.path.join(RECEIVED_PATH, fname)
        wb_r = load_workbook(fpath, data_only=True)

        if SHEET_NAME not in wb_r.sheetnames:
            continue

        ws_r = wb_r[SHEET_NAME]

        # Group rows from received file by Client_Code
        received_index = {}
        for r_row_idx in range(2, ws_r.max_row + 1):
            rval = ws_r.cell(row=r_row_idx, column=1).value  # Client_Code is column 1
            if rval:
                rkey = str(rval).strip()
                if rkey:
                    received_index.setdefault(rkey, []).append(r_row_idx)

        # Compare for each client
        for client_code, r_rows in received_index.items():
            master_rows = master_index.get(client_code, [])

            active_rows = [
                m_row
                for m_row in master_rows
                if str(
                    ws_master.cell(row=m_row, column=34).value or ""
                ).strip().lower()
                != "delete"
            ]

            if not active_rows:
                continue

            kept_row = active_rows[0]
            master_vals = get_row_values(ws_master, kept_row)

            for r_row in r_rows:
                recv_vals = get_row_values(ws_r, r_row)

                for idx in range(LAST_COL_INDEX):
                    if master_vals[idx] != recv_vals[idx]:
                        new_row = ws_change.max_row + 1
                        col_number = idx + 1

                        ws_change.cell(row=new_row, column=1).value = client_code
                        ws_change.cell(row=new_row, column=2).value = col_number
                        ws_change.cell(row=new_row, column=3).value = master_headers.get(
                            col_number, "Unknown"
                        )
                        ws_change.cell(row=new_row, column=4).value = master_vals[idx]
                        ws_change.cell(row=new_row, column=5).value = recv_vals[idx]

    # ---------- Apply formatting to Changed_Line_Items ----------
    thin_side = Side(border_style="thin", color="000000")
    all_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    header_font = Font(name="Aptos Narrow", size=10, bold=True, color="FFFFFF")
    data_font = Font(name="Aptos Narrow", size=10)

    fill_client = PatternFill(fill_type="solid", fgColor="7030A0")
    fill_other = PatternFill(fill_type="solid", fgColor="70A8E0")

    max_row = ws_change.max_row
    max_col = len(change_headers)

    for row in ws_change.iter_rows(
        min_row=1, max_row=max_row, min_col=1, max_col=max_col
    ):
        for cell in row:
            if cell.row == 1:
                cell.font = header_font
                cell.fill = fill_client if cell.column == 1 else fill_other
            else:
                cell.font = data_font

            cell.border = all_border
            cell.alignment = Alignment(
                vertical="top", horizontal="left", wrap_text=True
            )

    # Adjust column widths
    for col_idx in range(1, max_col + 1):
        col_letter = ws_change.cell(row=1, column=col_idx).column_letter
        if col_idx == 1:
            ws_change.column_dimensions[col_letter].width = 18
        elif col_idx == 3:
            ws_change.column_dimensions[col_letter].width = 25
        else:
            ws_change.column_dimensions[col_letter].width = 15

    # ---------- Save ----------
    wb_master.save(MASTER_PATH)
    print("Comparison complete → Changed items stored in 'Changed_Line_Items' sheet")


if __name__ == "__main__":
    ensure_backup(MASTER_PATH)
    run_old_vs_new()
