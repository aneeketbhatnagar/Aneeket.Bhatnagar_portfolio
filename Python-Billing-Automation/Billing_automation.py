# billing_automation_final.py
# Requires:
#    pip install openpyxl pywin32 tkcalendar

import os
import shutil
import datetime
import tkinter as tk
import OldVsNew  # must be present in same folder
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from copy import copy
import win32com.client as win32
import Filter_automation  # must be present in same folder
import EWS_automation  # must be present in same folder
from tkcalendar import DateEntry
from openpyxl.worksheet.filters import AutoFilter


# ======= CONFIG =======
# Use your real master path here. (Kept same as your original path)(https://clarivate.sharepoint.com/sites/BillingDashboards-DATABASE/Shared Documents/Master/Database FIleV1.xlsx)
MASTER_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Master\Database FIleV1.xlsx"
GENERATE_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Generate"
RECEIVED_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Received"

#MASTER_PATH = r"https://clarivate.sharepoint.com/sites/BillingDashboards-DATABASE/Shared Documents/Master/Database FIleV1.xlsx"
#GENERATE_PATH = r"https://clarivate.sharepoint.com/sites/BillingDashboards-DATABASE/Shared Documents/Generate"
#RECEIVED_PATH = r"https://clarivate.sharepoint.com/sites/BillingDashboards-DATABASE/Shared Documents/Received"
PROCESSED_PATH = os.path.join(RECEIVED_PATH, "Processed")

SHEET_NAME = "Monthly"

# ======= COLUMN INDEX CONFIG (1-based Excel column numbers) =======
COL_Client_Code = 1
COL_Manager = 2
COL_Billing_Entity = 3
COL_Lead = 4
COL_Customer_AS_PER_SOW = 5
COL_Order_Title_AS_per_SOW = 6
COL_Billing_Type = 7

# Editable columns (to copy from Received -> Master)
COL_No_of_FTEs = 8
COL_Fee_per_FTE = 9
COL_Total_FTE_Fee = 10
COL_Hours_Per_FTE = 11
COL_Total_Committed_Hours = 12
COL_Actual_Hours = 13
COL_Overage_Hours = 14
COL_Overage_Rate = 15
COL_Total_Overages = 16
COL_Total_Amount = 17
COL_Narrative = 24
COL_Comment_Or_Instructions = 29

# Billing_Status column index in both Received and Master
COL_Billing_Status = 31

# Status tracking column in Master (BillingPocStatus)
COL_BillingPocStatus = 34

# Last column index for formatting
LAST_COL_INDEX = 34

EDITABLE_COL_INDICES = [
    COL_No_of_FTEs, COL_Fee_per_FTE, COL_Total_FTE_Fee, COL_Hours_Per_FTE,
    COL_Total_Committed_Hours, COL_Actual_Hours, COL_Overage_Hours, COL_Overage_Rate,
    COL_Total_Overages, COL_Total_Amount, COL_Narrative, COL_Comment_Or_Instructions
]

# ======= HELPERS =======
def ensure_folders():
    os.makedirs(GENERATE_PATH, exist_ok=True)
    os.makedirs(RECEIVED_PATH, exist_ok=True)
    os.makedirs(PROCESSED_PATH, exist_ok=True)

def backup_master():
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    base, ext = os.path.splitext(MASTER_PATH)
    backup = f"{base}_backup_{ts}{ext}"
    shutil.copy2(MASTER_PATH, backup)
    return backup

def find_sheet_name_case_insensitive(wb, target):
    for s in wb.sheetnames:
        if s.strip().lower() == target.strip().lower():
            return s
    return None

def safe_trim(v):
    if v is None:
        return ""
    return str(v).strip()


# ======= FORMATTING =======
def apply_standard_formatting():
    try:
        if not os.path.exists(MASTER_PATH):
            messagebox.showerror("Formatting Error", f"Master file not found:\n{MASTER_PATH}")
            return
        wb = load_workbook(MASTER_PATH)
        sheets_to_format = ["Monthly", "Filtered_Data", "Blank_Status_Report"]
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        if "Monthly" not in wb.sheetnames:
            messagebox.showerror("Formatting Error", "'Monthly' sheet not found in workbook.")
            return
        src_ws = wb["Monthly"]
        header_styles = {}
        for col in range(1, LAST_COL_INDEX + 1):
            cell = src_ws.cell(row=1, column=col)
            header_styles[col] = {
                "font": copy(cell.font) if cell.font is not None else None,
                "fill": copy(cell.fill) if cell.fill is not None else None,
                "alignment": copy(cell.alignment) if cell.alignment is not None else None,
                "border": copy(cell.border) if cell.border is not None else None,
                "number_format": cell.number_format
            }
        for sheet_name in sheets_to_format:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            try:
                ws.sheet_view.showGridLines = False
            except Exception:
                pass
            max_row = ws.max_row if ws.max_row and ws.max_row > 1 else 1
            max_col = LAST_COL_INDEX
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                hs = header_styles.get(col)
                if hs:
                    try:
                        if hs["font"] is not None:
                            cell.font = copy(hs["font"])
                        if hs["fill"] is not None:
                            cell.fill = copy(hs["fill"])
                        if hs["alignment"] is not None:
                            cell.alignment = copy(hs["alignment"])
                        if hs["border"] is not None:
                            cell.border = copy(hs["border"])
                        cell.number_format = hs.get("number_format", cell.number_format)
                    except Exception:
                        pass
            for row in range(2, max_row + 1):
                ws.row_dimensions[row].height = 15
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    try:
                        cell.font = Font(name="Calibri", size=9)
                    except Exception:
                        pass
                    try:
                        cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                    except Exception:
                        pass
                    try:
                        cell.border = border
                    except Exception:
                        pass
        wb.save(MASTER_PATH)
    except Exception as e:
        messagebox.showerror("Formatting Error", str(e))

# ======= GENERATE MANAGER FILES (Excel COM) =======
def generate_manager_files():
    try:
        ensure_folders()
        selected_manager = manager_var.get().strip() if 'manager_var' in globals() else "All"
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        wb = excel.Workbooks.Open(MASTER_PATH)
        try:
            ws = wb.Sheets(SHEET_NAME)
        except Exception:
            sheet_real = None
            for s in wb.Sheets:
                if s.Name.strip().lower() == SHEET_NAME.strip().lower():
                    sheet_real = s
                    break
            if sheet_real is None:
                wb.Close(False)
                excel.Quit()
                messagebox.showerror("Generate Error", f"Sheet '{SHEET_NAME}' not found.")
                return
            ws = sheet_real
        header_row = 1
        max_col = ws.UsedRange.Columns.Count
        headers = [ws.Cells(header_row, c).Value for c in range(1, max_col + 1)]
        # find manager col position by header name if available, otherwise use COL_Manager
        if "Manager" in headers:
            manager_idx = headers.index("Manager") + 1
        else:
            manager_idx = COL_Manager
        last_row = ws.UsedRange.Rows.Count
        manager_rows = {}
        for r in range(2, last_row + 1):
            mgr = ws.Cells(r, manager_idx).Value
            if mgr:
                key = str(mgr).strip()
                manager_rows.setdefault(key, []).append(r)
        if selected_manager != "All":
            if selected_manager not in manager_rows:
                messagebox.showerror("Error", f"No records found for manager:\n{selected_manager}")
                wb.Close(False)
                excel.Quit()
                return
            manager_rows = {selected_manager: manager_rows[selected_manager]}
        for mgr, rows in manager_rows.items():
            new_wb = excel.Workbooks.Add()
            new_ws = new_wb.Sheets(1)
            new_ws.Name = SHEET_NAME
            ws.Rows(header_row).Copy(new_ws.Rows(1))
            target = 2
            for r in rows:
                ws.Rows(r).Copy(new_ws.Rows(target))
                target += 1
            safe = mgr.replace(" ", "_")
            path = os.path.join(GENERATE_PATH, f"{safe}_Monthly.xlsx")
            new_wb.SaveAs(path)
            new_wb.Close(SaveChanges=True)
        wb.Close(SaveChanges=False)
        excel.Quit()
        messagebox.showinfo("Success", f"Manager files generated in:\n{GENERATE_PATH}")
    except Exception as exc:
        try:
            excel.Quit()
        except Exception:
            pass
        messagebox.showerror("Generate error", str(exc))

# ======= UPDATE MASTER FROM RECEIVED (MAIN FUNCTION) =======
def update_master_from_received():
    try:
        ensure_folders()
        files = [f for f in os.listdir(RECEIVED_PATH) if f.lower().endswith((".xlsx", ".xlsm"))]
        if not files:
            messagebox.showinfo("Info", "No received files found.")
            return

        wb_master = load_workbook(MASTER_PATH, data_only=False)
        monthly_sheet_name = find_sheet_name_case_insensitive(wb_master, SHEET_NAME)
        if not monthly_sheet_name:
            messagebox.showerror("Error", f"Sheet '{SHEET_NAME}' not found in master.")
            return
        ws_master = wb_master[monthly_sheet_name]

        master_index = {}
        for r in range(2, ws_master.max_row + 1):
            v = ws_master.cell(row=r, column=COL_Client_Code).value
            if v is None:
                continue
            k = str(v).strip()
            master_index.setdefault(k, []).append(r)

        backup = backup_master()
        updates = 0
        added = 0
        deleted = 0

        for fname in files:
            fpath = os.path.join(RECEIVED_PATH, fname)
            wb_r = load_workbook(fpath, data_only=True)
            sheet_name_r = find_sheet_name_case_insensitive(wb_r, SHEET_NAME)
            if not sheet_name_r:
                shutil.move(fpath, os.path.join(PROCESSED_PATH, fname))
                continue
            ws_r = wb_r[sheet_name_r]

            received_groups = {}
            for r in range(2, ws_r.max_row + 1):
                v = ws_r.cell(row=r, column=COL_Client_Code).value
                if v is None:
                    continue
                k = str(v).strip()
                received_groups.setdefault(k, []).append(r)

            for client_code, r_rows in received_groups.items():
                new_count = len(r_rows)
                existing_master_rows = master_index.get(client_code, []).copy()
                old_count = len(existing_master_rows)

                # Transactional: multiple received rows
                if new_count > 1:
                    for mr in existing_master_rows:
                        ws_master.cell(row=mr, column=COL_BillingPocStatus).value = "Delete"
                        deleted += 1

                    for r in r_rows:
                        new_row = ws_master.max_row + 1
                        # copy all columns 1..LAST_COL_INDEX from received to master
                        for col_idx in range(1, LAST_COL_INDEX + 1):
                            ws_master.cell(row=new_row, column=col_idx).value = ws_r.cell(row=r, column=col_idx).value
                        ws_master.cell(row=new_row, column=COL_BillingPocStatus).value = "New"
                        added += 1

                    # update master_index to include appended rows
                    appended_range = list(range(ws_master.max_row - new_count + 1, ws_master.max_row + 1))
                    master_index[client_code] = existing_master_rows + appended_range
                    continue

                # Normal: single received row
                src_r = r_rows[0]

                # If no existing master row -> append new full row (but as per requirement, copy editable cols + Billing_Status + Client_Code)
                if old_count == 0:
                    new_row = ws_master.max_row + 1
                    # copy editable columns
                    for col_idx in EDITABLE_COL_INDICES:
                        ws_master.cell(row=new_row, column=col_idx).value = ws_r.cell(row=src_r, column=col_idx).value
                    # Billing_Status
                    ws_master.cell(row=new_row, column=COL_Billing_Status).value = ws_r.cell(row=src_r, column=COL_Billing_Status).value
                    # Client_Code
                    ws_master.cell(row=new_row, column=COL_Client_Code).value = ws_r.cell(row=src_r, column=COL_Client_Code).value
                    ws_master.cell(row=new_row, column=COL_BillingPocStatus).value = "New"
                    added += 1
                    master_index.setdefault(client_code, []).append(new_row)
                    continue

                # There are existing master rows -> find active (non-Delete)
                active_rows = []
                for mr in existing_master_rows:
                    st = ws_master.cell(row=mr, column=COL_BillingPocStatus).value
                    st_norm = "" if st is None else str(st).strip().lower()
                    if st_norm == "delete":
                        continue
                    active_rows.append(mr)

                # No active rows -> append new
                if len(active_rows) == 0:
                    new_row = ws_master.max_row + 1
                    for col_idx in EDITABLE_COL_INDICES:
                        ws_master.cell(row=new_row, column=col_idx).value = ws_r.cell(row=src_r, column=col_idx).value
                    ws_master.cell(row=new_row, column=COL_Billing_Status).value = ws_r.cell(row=src_r, column=COL_Billing_Status).value
                    ws_master.cell(row=new_row, column=COL_Client_Code).value = ws_r.cell(row=src_r, column=COL_Client_Code).value
                    ws_master.cell(row=new_row, column=COL_BillingPocStatus).value = "New"
                    added += 1
                    master_index.setdefault(client_code, []).append(new_row)
                    continue

                # If multiple active rows -> keep first, mark others Delete
                kept_row = active_rows[0]
                if len(active_rows) > 1:
                    for extra in active_rows[1:]:
                        ws_master.cell(row=extra, column=COL_BillingPocStatus).value = "Delete"
                        deleted += 1

                # Update editable fields on kept row
                for col_idx in EDITABLE_COL_INDICES:
                    ws_master.cell(row=kept_row, column=col_idx).value = ws_r.cell(row=src_r, column=col_idx).value

                # Overwrite Billing_Status in master with received Billing_Status
                ws_master.cell(row=kept_row, column=COL_Billing_Status).value = ws_r.cell(row=src_r, column=COL_Billing_Status).value

                # BillingPocStatus progression (based only on MASTER current value)
                current = ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value
                curr_norm = "" if current is None else str(current).strip().lower()

                if curr_norm == "":
                    ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value = "New"
                elif curr_norm == "new":
                    ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value = "Updated"
                elif curr_norm == "updated":
                    ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value = "Updated2"
                elif curr_norm == "updated2":
                    ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value = "Updated3"
                else:
                    # leave other values unchanged
                    ws_master.cell(row=kept_row, column=COL_BillingPocStatus).value = current

                updates += 1

            # move processed file
            shutil.move(fpath, os.path.join(PROCESSED_PATH, fname))

        wb_master.save(MASTER_PATH)
        apply_standard_formatting()

        messagebox.showinfo(
            "Success",
            f"Master Updated (Normal Updates): {updates}\n"
            f"New Rows Added                : {added}\n"
            f"Old Rows Marked DELETE        : {deleted}\n\n"
            f"Backup saved: {backup}"
        )

    except Exception as e:
        messagebox.showerror("Update Error", str(e))

# ======= RUN FILTER AUTOMATION (Date picker integrated) =======
def run_filter_auto():
    try:
        # DateEntry returns a datetime.date object; we will format as YYYY-MMM-DD (example: 2025-Jan-15)
        selected_date_obj = date_picker.get_date()
        selected_date_str = selected_date_obj.strftime("%Y-%b-%d")  # YYYY-MMM-DD as requested
        # Pass string to your Filter_automation function (assumed to accept string)
        Filter_automation.create_filtered_sheet(selected_date_str)
        apply_standard_formatting()
        messagebox.showinfo("Success", f"Filter Automation completed for: {selected_date_str}")
    except Exception as e:
        messagebox.showerror("Filter Error", str(e))

# ======= RUN EWS FILTER AUTOMATION (Date picker integrated) =======
def run_ewsreport_auto(*args):
    try:
        selected = date_picker.get_date()
        # Fake calendar object
        fake_cal = lambda: None
        fake_cal.get_date = lambda: selected
        EWS_automation.cal = fake_cal
        EWS_automation.generate_ews_report()
        messagebox.showinfo("EWS SUCCESS", "EWS Report Generated & Opened!")
    except Exception as e:
        messagebox.showerror("EWS Error", str(e))


# ======= Email and Blank report =======
def send_blank_status_email():
    try:
        if not os.path.exists(MASTER_PATH):
            messagebox.showerror("Error", f"Master file not found:\n{MASTER_PATH}")
            return
        wb = load_workbook(MASTER_PATH, data_only=True)
        monthly_sheet = find_sheet_name_case_insensitive(wb, SHEET_NAME)
        if not monthly_sheet:
            messagebox.showerror("Error", f"Sheet '{SHEET_NAME}' not found.")
            return
        ws = wb[monthly_sheet]
        data = []
        for r in range(2, ws.max_row + 1):
            status = ws.cell(r, 34).value            #status = ws.cell(r, COL_Billing_Status).value
            if status is None or str(status).strip() == "":
                client = ws.cell(r, COL_Client_Code).value or ""
                manager = ws.cell(r, COL_Manager).value or "N/A"
                sr_manager = ws.cell(r, COL_Customer_AS_PER_SOW).value or "N/A"
                lead = ws.cell(r, COL_Lead).value or "N/A"
                data.append((client, manager, sr_manager, lead))
        if not data:
            messagebox.showinfo("Done", "No blank Billing_Status found.")
            return
        html_body = """
        <html><body style="font-family: Calibri; font-size: 12pt;">
        <p>Dear Team,</p>
        <p>Requesting billing inputs for the clients listed below.</p>
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; font-size: 11pt;">
            <tr style="background-color: #154360; color: white; font-weight: bold;">
                <td>Sr</td><td>Client Code</td><td>Client Name</td><td>Manager</td><td>Lead</td>
            </tr>
        """
        for i, row in enumerate(data, start=1):
            html_body += f"<tr><td>{i}</td><td>{row[0]}</td><td>{row[2]}</td><td>{row[1]}</td><td>{row[3]}</td></tr>"
        html_body += "</table><br><p>Regards,<br>Aneeket</p></body></html>"
        outlook = win32.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.Subject = "Billing Input Pending – Action Required"
        mail.HTMLBody = html_body
        mail.Display()
        messagebox.showinfo("Email Created", "Email ready in Outlook.")
    except Exception as e:
        messagebox.showerror("Email Error", str(e))

def generate_blank_billing_status_report():
    try:
        if not os.path.exists(MASTER_PATH):
            messagebox.showerror("Error", f"Master file not found:\n{MASTER_PATH}")
            return
        wb = load_workbook(MASTER_PATH)
        monthly_sheet = find_sheet_name_case_insensitive(wb, SHEET_NAME)
        if not monthly_sheet:
            messagebox.showerror("Error", f"Sheet '{SHEET_NAME}' not found.")
            return
        ws = wb[monthly_sheet]
        if "Blank_Status_Report" in wb.sheetnames:
            del wb["Blank_Status_Report"]
        report_ws = wb.create_sheet("Blank_Status_Report")
        for c in range(1, ws.max_column + 1):
            report_ws.cell(1, c).value = ws.cell(1, c).value
        target_row = 2
        for r in range(2, ws.max_row + 1):
            status = ws.cell(r, COL_BillingPocStatus).value
            if status is None or str(status).strip() == "":
                for c in range(1, ws.max_column + 1):
                    report_ws.cell(target_row, c).value = ws.cell(r, c).value
                target_row += 1
        wb.save(MASTER_PATH)
        apply_standard_formatting()
        messagebox.showinfo("Report Created", f"Report created in sheet 'Blank_Status_Report'\nTotal rows: {target_row-2}")
    except Exception as e:
        messagebox.showerror("Error", str(e))


def run_oldvsnew_from_gui():
    try:
        OldVsNew.run_old_vs_new()
        messagebox.showinfo(
            "Success",
            "Old vs New comparison completed.\nChanged_Line_Items sheet updated."
        )
    except Exception as e:
        messagebox.showerror("Comparison Error", str(e))


# ======= GUI =======
BG = "#E9F5FF"
PRIMARY = "#356EFF"
P_HOVER = "#3658FF"
WHITE = "#FFFFFF"
SECOND = "#FFFFFF"
S_HOVER = "#F1F1F1"
EXIT_BG = "#E63946"
EXIT_HOVER = "#C62828"

def make_btn(text, bg, fg, hover, command):
    btn = tk.Label(root, text=text, bg=bg, fg=fg,
                   font=("Segoe UI", 10, "bold"),
                   bd=1, relief="solid", width=32, height=2)
    btn.bind("<Button-1>", lambda e: command())
    btn.bind("<Enter>", lambda e: btn.config(bg=hover))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn

def exit_app():
    root.destroy()

if __name__ == "__main__":
    ensure_folders()
    root = tk.Tk()
    root.title("Billing Automation Tool")
    root.geometry("520x720")
    root.configure(bg=BG)
    root.resizable(False, False)

    title = tk.Label(root, text="Billing Automation Tool", font=("Segoe UI", 22, "bold"), fg="black", bg=BG)
    title.pack(pady=12)

    # Date picker (Full date) - user requested format: YYYY-MMM-DD (we will format when calling filter)
    frame = tk.Frame(root, bg=BG)
    frame.pack(pady=8)
    tk.Label(frame, text="Select Date (Full):", font=("Segoe UI", 10, "bold"), bg=BG).pack(side="left", padx=8)
    date_picker = DateEntry(frame, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='y-mm-dd')
    date_picker.pack(side="left")

    btn_filter = make_btn("Run Filter Automation (by Date)", PRIMARY, WHITE, P_HOVER, run_filter_auto)
    btn_filter.pack(pady=10)

    # manager dropdown loader
    def load_managers_for_dropdown():
        try:
            wb = load_workbook(MASTER_PATH, data_only=True)
            sheet = find_sheet_name_case_insensitive(wb, SHEET_NAME)
            if not sheet:
                return ["All"]
            ws = wb[sheet]
            mgrs = set()
            for r in range(2, ws.max_row + 1):
                v = ws.cell(row=r, column=COL_Manager).value
                if v:
                    mgrs.add(str(v).strip())
            return ["All"] + sorted(list(mgrs))
        except Exception:
            return ["All"]

    manager_list = load_managers_for_dropdown()
    manager_var = tk.StringVar()
    manager_var.set("All")
    mgr_frame = tk.Frame(root, bg=BG)
    mgr_frame.pack(pady=6)
    tk.Label(mgr_frame, text="Select Manager:", font=("Segoe UI", 10, "bold"), bg=BG).pack(side="left", padx=8)
    mgr_box = tk.OptionMenu(mgr_frame, manager_var, *manager_list)
    mgr_box.config(font=("Segoe UI", 10), bg=WHITE, fg="black", width=14, anchor="w")
    mgr_box.pack(side="left")

    btn2 = make_btn("Generate Manager Files", PRIMARY, WHITE, P_HOVER, generate_manager_files)
    btn2.pack(pady=8)
    btn3 = make_btn("Update Master from Received Files", PRIMARY, WHITE, P_HOVER, update_master_from_received)
    btn3.pack(pady=8)
    btn4 = make_btn("Blank Billing Status Report", PRIMARY, WHITE, P_HOVER, generate_blank_billing_status_report)
    btn4.pack(pady=8)
    btn_email = make_btn("Email: Blank Billing Status", PRIMARY, WHITE, P_HOVER, send_blank_status_email)
    btn_email.pack(pady=8)
    btn_old_new = make_btn("Run Old vs New Compare", PRIMARY, WHITE, P_HOVER, run_oldvsnew_from_gui)
    btn_old_new.pack(pady=8)
    btn_filter = make_btn("EWS Automation (by Date)", PRIMARY, WHITE, P_HOVER, run_ewsreport_auto)
    btn_filter.pack(pady=8)

    btn_format = make_btn("Apply Formatting to All Sheets", SECOND, 'black', S_HOVER, apply_standard_formatting)
    btn_format.pack(pady=8)
    exit_btn = make_btn("Exit", EXIT_BG, WHITE, EXIT_HOVER, exit_app)
    exit_btn.pack(pady=12)

    root.mainloop()
