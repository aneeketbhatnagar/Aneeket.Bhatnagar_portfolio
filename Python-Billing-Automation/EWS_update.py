# Update_From_EWS.py
# BOSS KA FINAL ORDER - Ek Client Ek Baar | All services same data

import openpyxl
import os
import tkinter as tk
from tkinter import filedialog, messagebox

MASTER_PATH = r"C:\Users\U6028093\OneDrive - Clarivate Analytics\Desktop\Backup\office\Billing Data\2026\Master\Database FIleV1.xlsx"

def get_client_prefix(client_code):
    """CPAG-0001-00 → CPAG-0001 | CPAG-0001-01 → CPAG-0001"""
    if not client_code or '-' not in client_code:
        return client_code
    parts = client_code.split('-')
    if len(parts) >= 3:
        return '-'.join(parts[:2])  # CPAG-0001
    return client_code

def update_from_ews():
    try:
        root = tk.Tk()
        root.withdraw()
        ews_path = filedialog.askopenfilename(
            title="Select EWS Report (One Client Per Row)",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not ews_path:
            return

        # Load files
        wb_master = openpyxl.load_workbook(MASTER_PATH)
        ws_master = wb_master["Master_Database"]
        ws_monthly = wb_master["Monthly_Master_Database"]
        wb_ews = openpyxl.load_workbook(ews_path)
        ws_ews = wb_ews.active

        # EWS headers
        ews_headers = {str(ws_ews.cell(1,c).value or "").strip(): c for c in range(1, ws_ews.max_column + 1)}

        # Last RAG column
        rag_columns = [k for k in ews_headers if k.endswith("_RAG")]
        if not rag_columns:
            messagebox.showerror("Error", "No RAG columns!")
            return
        latest_rag_col_name = rag_columns[-1]
        latest_rag_col = ews_headers[latest_rag_col_name]
        month_key = latest_rag_col_name.replace("_RAG", "")
        target_month_year = f"20{month_key[3:5]}-{month_key[:3]}"  # Nov25 → 2025-Nov

        # Master headers
        master_headers = {}
        for c in range(1, ws_master.max_column + 1):
            val = str(ws_master.cell(1, c).value or "").strip()
            if val:
                master_headers[val] = c
                master_headers[val.replace(" ", "")] = c

        # Monthly columns
        client_col_m = rag_col_m = month_col_m = None
        for c in range(1, ws_monthly.max_column + 1):
            h = str(ws_monthly.cell(1, c).value or "").strip()
            if "Client_Code" in h: client_col_m = c
            if "RAG_Status" in h: rag_col_m = c
            if "Billing_Month" in h: month_col_m = c

        if not all([client_col_m, rag_col_m, month_col_m]):
            messagebox.showerror("Error", "Monthly columns missing!")
            return

        updated_rows = 0

        # Loop through EWS (one row per client)
        for r in range(2, ws_ews.max_row + 1):
            ews_client_code = str(ws_ews.cell(r, ews_headers["Client_Code"]).value or "").strip()
            if not ews_client_code: continue

            client_prefix = get_client_prefix(ews_client_code)  # CPAG-0001

            # Get latest values from EWS
            update_values = {}
            columns_to_sync = [
                "Client_POC", "CSM_Contact", "EWS_Indicator",
                "RAG_Historic_Comments", "RAG_Current_Brief_Summary"
            ]
            for col_name in columns_to_sync:
                if col_name in ews_headers:
                    update_values[col_name] = ws_ews.cell(r, ews_headers[col_name]).value

            rag_val = str(ws_ews.cell(r, latest_rag_col).value or "").strip().upper()

            # === Update ALL service rows with this client prefix ===
            for mr in range(2, ws_master.max_row + 1):
                master_client = str(ws_master.cell(mr, master_headers["Client_Code"]).value or "").strip()
                if get_client_prefix(master_client) == client_prefix:
                    # Update Master_Database columns
                    for col_name, val in update_values.items():
                        dest_col = master_headers.get(col_name) or master_headers.get(col_name.replace("_", " "))
                        if dest_col:
                            ws_master.cell(mr, dest_col).value = val

                    updated_rows += 1

            # === Update Monthly_Master_Database - All services same RAG ===
            if rag_val:
                for mr in range(2, ws_monthly.max_row + 1):
                    m_client = str(ws_monthly.cell(mr, client_col_m).value or "").strip()
                    m_month = str(ws_monthly.cell(mr, month_col_m).value or "").strip()
                    if get_client_prefix(m_client) == client_prefix and target_month_year in m_month:
                        ws_monthly.cell(mr, rag_col_m).value = rag_val

        # Save
        wb_master.save(MASTER_PATH)
        os.startfile(MASTER_PATH)

        messagebox.showinfo("BOSS KHUSH HO GAYA!",
                            f"Update Complete!\n\n"
                            f"{updated_rows} service rows updated\n"
                            f"Client Grouping: CPAG-0001, CPAG-0002, etc.\n"
                            f"Latest RAG: {latest_rag_col_name}\n"
                            f"File Opened!")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# ===================== RUN =====================
if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    update_from_ews()